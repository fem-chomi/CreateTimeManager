import openpyxl
import json

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from datetime import datetime
from datetime import timedelta


MAX_ROWS = 200


def CreateCommonCell(ws, row, col):
	cell = ws.cell(row = row, column = col)
	setAlignmentCenter(cell)
	setBorder(cell)
	return cell
	

def setAlignmentCenter(cell):
  cell.alignment = Alignment(horizontal="center", vertical="center")


def setBorder(cell):
  side = Side(style='thin', color='000000')
  border = Border(top=side, bottom=side, left=side, right=side)
  cell.border = border


# 左上エリア
def CreateHeader(ws):
	row = 1
	col = 1
	cell = CreateCommonCell(ws, row, col)
	cell.value = 'Please write by the day before! (UTC+9)'
	cell.font = Font(color='dc143c', bold=True)
	ws.merge_cells('A1:C1')
	
	row = 2
	col = 1
	cell = CreateCommonCell(ws, row, col)
	cell.value = 'Number of People'
	ws.merge_cells('A2:C2')
	
	row = 3
	col = 1
	count = int(setting['team_max_count'])
	cell = CreateCommonCell(ws, row, col)
	cell.value = f'Number of Units ({count}x{count})'
	ws.merge_cells('A3:C3')
	
	# current datetime
	row = 4
	col = 1
	cell = CreateCommonCell(ws, row, col)
	cell.value = '=now()'
	cell.number_format = 'yyyy/mm/dd hh:mm:ss'
	cell.font = Font(color='dc143c', bold=True)
	
	# event url
	row = 4
	col = 2
	cell = CreateCommonCell(ws, row, col)
	cell.value = setting['event_name']
	cell.hyperlink = setting['event_url']
	cell.font = Font(color='4169e1', bold=True)
	ws.merge_cells('B4:C4')
	

# 右上エリア
def CreateEventSchedule(ws, setting):
  col = 4
  s_format = '%Y/%m/%d'
  begin_date = datetime.strptime(setting['event_begin_date'], s_format)
  end_date = datetime.strptime(setting['event_end_date'], s_format)
  for n in range((end_date - begin_date).days + 1):
    current_date = begin_date + timedelta(n)
    for time in setting['event_times']:
      ws.column_dimensions[get_column_letter(col)].width = 12
      letter = get_column_letter(col)

      # number of people
      row = 2
      cell = CreateCommonCell(ws, row, col)
      cell.value = f'=COUNTIF({letter}6:{letter}{MAX_ROWS-1}, "○")'

      # number of Units
      row = 3
      cell = CreateCommonCell(ws, row, col)
      count = int(setting['team_max_count'])
      if count > 0:
        #cell.value = f'={letter}2/{count})'
        cell.value = f'=INT({letter}2/{count})'
	
      # date
      row = 4
      cell = CreateCommonCell(ws, row, col)
      cell.value = f'{current_date.month}/{current_date.day}({current_date.strftime("%a")})'
      cell.number_format = 'm/d'
      if current_date.strftime("%a") == 'Sat':
        cell.font = Font(color='4169e1', bold=True)
      elif current_date.strftime("%a") == 'Sun':
        cell.font = Font(color='dc143c', bold=True)
      else:
        cell.font = Font(bold=True)
	
      # time
      row = 5
      cell = CreateCommonCell(ws, row, col)
      cell.value = time
      cell.font = Font(bold=True)

      col += 1


# 左下エリア
def CreateMemberList(ws):
	row = 5
	col = 1
	cell = CreateCommonCell(ws, row, col)
	cell.value = 'In Game Name'
	cell.font = Font(bold=True)
	ws.column_dimensions['A'].width = 20
	
	row = 5
	col = 2
	cell = CreateCommonCell(ws, row, col)
	cell.value = 'Discord ID'
	cell.font = Font(bold=True)
	ws.column_dimensions['B'].width = 20
	
	row = 5
	col = 3
	cell = CreateCommonCell(ws, row, col)
	cell.value = 'Language'
	cell.font = Font(bold=True)
	ws.column_dimensions['C'].width = 10
	
	row = 6
	for member in setting['members']:
		col = 1
		cell = CreateCommonCell(ws, row, col)
		cell.value = member['IGN']
		
		col = 2
		cell = CreateCommonCell(ws, row, col)
		cell.value = member['DiscordName']
		
		col = 3
		cell = CreateCommonCell(ws, row, col)
		cell.value = member['Language']
		row += 1

	for row in range(row, MAX_ROWS):
	  for col in range(1, 4):
	    cell = CreateCommonCell(ws, row, col)
	    cell.value = ''


# 右下エリア
def CreateMemberSchedule(ws):
  # schedules
  col = 4
  s_format = '%Y/%m/%d'
  begin_date = datetime.strptime(setting['event_begin_date'], s_format)
  end_date = datetime.strptime(setting['event_end_date'], s_format)
  for n in range((end_date - begin_date).days + 1):
    current_date = begin_date + timedelta(n)
    for time in setting['event_times']:
      letter = get_column_letter(col)
      row = 6
      for member in setting['members']:
        cell = CreateCommonCell(ws, row, col)
        dv = DataValidation(type="list", formula1='"{}"'.format(','.join(setting['selection_values'])), allow_blank=False, showErrorMessage=True, errorStyle="warning")
        dv.add(ws[f'{letter}{row}'])
        ws.add_data_validation(dv)
        row += 1

      for row in range(row, MAX_ROWS):
        cell = CreateCommonCell(ws, row, col)
        dv = DataValidation(type="list", formula1 = '"{}"'.format(','.join(setting['selection_values'])), allow_blank=False, showErrorMessage=True, errorStyle="warning")
        dv.add(ws[f'{letter}{row}'])
        ws.add_data_validation(dv)

      col += 1
	

wb = openpyxl.Workbook()
try:
    wb.remove(wb['Sheet'])
except:
    pass
ws = wb.create_sheet(title = 'IGN&Schedule')

setting = json.load(open("setting.json", "r", encoding="utf-8"))

CreateHeader(ws)
CreateEventSchedule(ws, setting)
CreateMemberList(ws)
CreateMemberSchedule(ws)

wb.save('time_managrer_.xlsx')
wb.close()

